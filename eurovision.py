import numpy as np
from eurovision_voting import eurovision_instance_opt_voting

import openpyxl
from pathlib import Path

xlsx_file = Path('.', 'eurovision_song_contest_1975_2019.xlsx')
wb_obj = openpyxl.load_workbook(xlsx_file)

# Read the active sheet:
sheet = wb_obj.active


# Select the year of the competition
year = 2004

countries_ids = {}
current_id = 0

for row in sheet.iter_rows():
    y = row[0].value
    if (y != year):
        continue
    stage = row[1].value

    # We are only interested in the final stage

    if (stage != 'f'):
        continue

    from_country = row[4].value

    if (from_country not in countries_ids):
        countries_ids[from_country] = current_id
        current_id += 1

n = len(countries_ids) # Number of voters

contesters_ids = []

for row in sheet.iter_rows():
    y = row[0].value
    if (y != year):
        continue
    stage = row[1].value
    if (stage != 'f'):
        continue

    to_country = row[5].value
    country_id = countries_ids[to_country]

    if (country_id not in contesters_ids):
        contesters_ids.append(country_id)

m = len(contesters_ids) # Number of candidates

# Define a mapping from scores to orders

orders = {
    12 : 0,
    10 : 1,
    8 : 2,
    7 : 3,
    6 : 4,
    5 : 5,
    4 : 6,
    3 : 7,
    2 : 8,
    1 : 9
}

countries_prefs = {}
countries_scores = {i : 0 for i in contesters_ids}

for i in range(n):
    countries_prefs[i] = [0 for i in range(10)]

for row in sheet.iter_rows():
    y = row[0].value
    if (y != year):
        continue
    stage = row[1].value
    if (stage != 'f'):
        continue

    from_country_id = countries_ids[row[4].value]
    to_country_id = countries_ids[row[5].value]
    points = row[6].value

    if (points == 0):
        continue

    countries_scores[to_country_id] += points

    order = orders[points]

    countries_prefs[from_country_id][order] = to_country_id

prefs = []

for i in range(n):

    p = countries_prefs[i]
    pref = []

    for j in range(len(p)):
        for r in contesters_ids:
            if (r in p[0:(j+1)]):
                continue
            if (r == i):
                continue
            pref.append((p[j], r))

    prefs.append(pref)

print(contesters_ids)

print(prefs[0])

#distortions = eurovision_instance_opt_voting(n, m, prefs, contesters_ids)
#print(distortions)
